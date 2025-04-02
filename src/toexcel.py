import os
import json
import pandas as pd
from openpyxl import load_workbook

json_folders = ["../dist/finished", "../dist/admin"]  # 두 개의 폴더에서 검색

all_data = []

# 🔍 모든 폴더에서 JSON 파일 찾기
for folder in json_folders:
    for root, _, files in os.walk(folder):  # 하위 폴더까지 탐색
        for filename in files:
            if filename.endswith(".json"):  # JSON 파일만 선택
                file_path = os.path.join(root, filename)

                # JSON 파일 읽기
                with open(file_path, "r", encoding="utf-8") as f:
                    data = json.load(f)

                    if isinstance(data, list):  # ✅ 리스트 형태인지 확인
                        for item in data:
                            if isinstance(item, dict):  # ✅ 각 항목이 딕셔너리인지 확인
                                all_data.append(item)
                    elif isinstance(data, dict):  # ✅ 단일 JSON 객체도 추가
                        all_data.append(data)

# 📊 JSON 데이터를 DataFrame으로 변환
df = pd.DataFrame(all_data)
df["img2Url"] = df.get("img2Url", "")
df["img3Url"] = df.get("img3Url", "")

print(df.head())  # 상위 5개 데이터만 출력
print(df.columns)  # 컬럼명이 정확한지 확인

# ✅ 엑셀 파일 저장 경로
excel_path = "../dist/excel/merged_data.xlsx"

# ✅ DataFrame을 엑셀 파일로 저장
df.to_excel(excel_path, index=False, engine="openpyxl")

# ✅ 강제로 특정 열에 배치할 딕셔너리
headers = {
    "B": "modelName",
    "F": "name",
    "N": "price",
    "I": "categoryCode",
    "S": "img1Url",
    "T": "img2Url",
    "U": "img3Url",
}

# ✅ 기존 엑셀 파일 열기
wb = load_workbook(excel_path)
ws = wb.active

# 🔄 새로운 시트 만들기
new_ws = wb.create_sheet(title="정렬된 데이터")

# 📌 헤더 추가
for col, header in headers.items():
    new_ws[f"{col}1"] = header  # 첫 번째 행에 헤더 넣기

# 📌 데이터 추가 (2번째 행부터 데이터 입력)
for row_idx, row in enumerate(df.itertuples(index=False), start=2):
    for col, header in headers.items():
        if header in df.columns:
            value = getattr(row, header, "")  # ✅ 수정된 부분 (값을 정확하게 가져옴)
            new_ws[f"{col}{row_idx}"] = value  # 해당 열(row_idx)에 값 추가

# 📁 변경된 데이터 저장
wb.remove(ws)  # 기존 시트 삭제
wb.save(excel_path)
print(f"✅ 엑셀 파일 저장 완료! '{excel_path}'")
